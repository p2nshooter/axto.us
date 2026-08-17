#!/usr/bin/env python3
"""
Menyusun berkas Daftar Pemilih Sementara (DPS) Desa Sukakarya
Pemilihan Kepala Desa Tahun 2026 -- lengkap dengan halaman COVER,
kop/judul resmi, dan satu lembar kerja per TPS.

Sumber data : Bagi_A._Daftar_Pemilih_DPSHP.xlsx (sheet "Olah Data")
"""

import datetime
import shutil
from collections import defaultdict

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break

SRC = "source.xlsx"
OUT = "DPS_Desa_Sukakarya_2026_per_TPS.xlsx"

# ---------------------------------------------------------------- identitas
DESA = "SUKAKARYA"
KECAMATAN = "SUKAKARYA"
KABUPATEN = "BEKASI"
TAHUN = "2026"
TGL_PEMUNGUTAN = datetime.datetime(2026, 9, 21)

# ---------------------------------------------------------------- palet warna
HIJAU = "1E4620"        # hijau tua - judul & kop
HIJAU_MUDA = "DCE7DC"   # arsir header tabel
EMAS = "A67C00"         # aksen emas
ABU = "F2F2F2"
PUTIH = "FFFFFF"

FONT = "Arial"

thin = Side(style="thin", color="808080")
medium = Side(style="medium", color="404040")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)
BOX_TEBAL = Border(left=medium, right=medium, top=medium, bottom=medium)

# kode disabilitas (sesuai legenda pada berkas asli, kolom AA:AC)
DISABILITAS = {
    1: "Fisik",
    2: "Intelektual",
    3: "Mental",
    4: "Tuna Wicara",
    5: "Tuna Rungu",
    6: "Tuna Netra",
}

# kolom tabel DPS
KOLOM = [
    ("NO", 6),
    ("NOMOR KARTU KELUARGA", 21),
    ("NIK", 21),
    ("NAMA LENGKAP", 32),
    ("TEMPAT LAHIR", 18),
    ("TANGGAL LAHIR", 13),
    ("STATUS PERKAWINAN\nB/S/P", 11),
    ("JENIS KELAMIN\nL/P", 9),
    ("JALAN/DUKUH", 20),
    ("RT", 5),
    ("RW", 5),
    ("KETERANGAN", 16),
]
NCOL = len(KOLOM)
LAST_COL = get_column_letter(NCOL)


# ============================================================ baca data sumber
def baca_sumber():
    wb = openpyxl.load_workbook(SRC)
    ws = wb["Olah Data"]
    data = []
    for r in range(3, ws.max_row + 1):
        nama = ws.cell(r, 4).value
        if not nama:
            continue
        data.append(
            {
                "row": r,
                "nkk": str(ws.cell(r, 2).value).strip(),
                "nik": str(ws.cell(r, 3).value).strip(),
                "nama": str(nama).strip(),
                "tempat": str(ws.cell(r, 5).value).strip().upper(),
                "tgl": ws.cell(r, 6).value,
                "kawin": str(ws.cell(r, 8).value).strip().upper(),
                "jk": str(ws.cell(r, 9).value).strip().upper(),
                # normalisasi: beberapa alamat pada berkas asli berspasi ganda/ekor
                "alamat": " ".join(str(ws.cell(r, 10).value).split()).upper(),
                "rt": int(ws.cell(r, 11).value or 0),
                "rw": int(ws.cell(r, 12).value or 0),
                "disab": int(ws.cell(r, 13).value or 0),
                # normalisasi huruf besar/kecil pada kolom E-KTP
                "ektp": str(ws.cell(r, 14).value or "").strip().upper(),
                "tps": int(ws.cell(r, 15).value or 0),
                "ketua_rt": str(ws.cell(r, 18).value or "").strip().upper(),
            }
        )
    return data


def perbaiki_olah_data(wb):
    """Perbaiki rumus DATEDIF pada sheet sumber.

    Berkas asli memakai =DATEDIF(F3,"21/09/2026","Y") -- tanggal ditulis
    sebagai TEKS, sehingga Excel mengembalikan #VALUE! (nilai #VALUE! itu
    memang sudah tersimpan di berkas asli, bukan akibat pengolahan ini).
    Diganti dengan DATE(2026,9,21) agar terhitung benar.
    """
    ws = wb["Olah Data"]
    diperbaiki = 0
    for r in range(3, ws.max_row + 1):
        for c, satuan in ((23, "Y"), (24, "YM"), (25, "MD")):
            cell = ws.cell(r, c)
            if isinstance(cell.value, str) and cell.value.startswith("=DATEDIF("):
                cell.value = (
                    f'=IF(F{r}="","",DATEDIF(F{r},DATE({TGL_PEMUNGUTAN.year},'
                    f'{TGL_PEMUNGUTAN.month},{TGL_PEMUNGUTAN.day}),"{satuan}"))'
                )
                diperbaiki += 1
    return diperbaiki


def usia_pada(tgl, acuan=TGL_PEMUNGUTAN):
    """Usia penuh (tahun) pada tanggal pemungutan suara."""
    if not isinstance(tgl, datetime.datetime):
        return None
    t = acuan.year - tgl.year
    if (acuan.month, acuan.day) < (tgl.month, tgl.day):
        t -= 1
    return t


# ============================================================ util penulisan
def tulis(ws, row, col, value, *, font=None, align=None, border=None,
          fill=None, fmt=None):
    c = ws.cell(row, col, value)
    c.font = font or Font(name=FONT, size=10)
    if align:
        c.alignment = align
    if border:
        c.border = border
    if fill:
        c.fill = fill
    if fmt:
        c.number_format = fmt
    return c


TENGAH = Alignment(horizontal="center", vertical="center", wrap_text=True)
KIRI = Alignment(horizontal="left", vertical="center")
TENGAH_ATAS = Alignment(horizontal="center", vertical="center")


# ============================================================ halaman COVER
def buat_cover(wb, per_tps, tps=None):
    """Halaman sampul.

    tps=None  -> sampul buku lengkap, memuat rekap seluruh TPS.
    tps=<int> -> sampul berkas satu TPS (dipakai untuk keluaran PDF per TPS).
    """
    ws = wb.create_sheet("COVER")
    ws.sheet_view.showGridLines = False

    for col, lebar in zip("ABCDEFGH", (3, 14, 14, 14, 14, 14, 14, 3)):
        ws.column_dimensions[col].width = lebar

    # bingkai luar hijau (persegi penuh, baris 2..45 kolom B..G)
    R1, R2, C1, C2 = 2, 45, 2, 7
    tepi = Side(style="medium", color=HIJAU)
    for r in range(R1, R2 + 1):
        for c in range(C1, C2 + 1):
            ws.cell(r, c).fill = PatternFill("solid", fgColor=PUTIH)
            ws.cell(r, c).border = Border(
                left=tepi if c == C1 else None,
                right=tepi if c == C2 else None,
                top=tepi if r == R1 else None,
                bottom=tepi if r == R2 else None,
            )

    def baris(r, teks, size, color=HIJAU, bold=True, tinggi=None, italic=False):
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        tulis(ws, r, 2, teks,
              font=Font(name=FONT, size=size, bold=bold, italic=italic, color=color),
              align=TENGAH_ATAS)
        ws.row_dimensions[r].height = tinggi or (size * 1.9)

    # garis emas atas
    for c in range(2, 8):
        ws.cell(4, c).fill = PatternFill("solid", fgColor=EMAS)
    ws.row_dimensions[4].height = 6

    baris(7, "DAFTAR PEMILIH SEMENTARA", 30, tinggi=44)
    baris(9, "( D P S )", 22, color=EMAS, tinggi=32)

    for c in range(2, 8):
        ws.cell(11, c).fill = PatternFill("solid", fgColor=EMAS)
    ws.row_dimensions[11].height = 4

    baris(14, f"DESA {DESA}", 24, tinggi=36)
    baris(15, f"KECAMATAN {KECAMATAN}", 15, tinggi=24)
    baris(16, f"KABUPATEN {KABUPATEN}", 15, tinggi=24)

    # penanda TPS -- hanya pada sampul berkas per TPS
    if tps is not None:
        ws.merge_cells(start_row=18, start_column=3, end_row=18, end_column=6)
        tulis(ws, 18, 3, f"TEMPAT PEMUNGUTAN SUARA (TPS)  {tps:02d}",
              font=Font(name=FONT, size=18, bold=True, color=PUTIH),
              align=TENGAH_ATAS,
              fill=PatternFill("solid", fgColor=HIJAU))
        ws.row_dimensions[18].height = 34

    baris(20, "PANITIA PEMILIHAN KEPALA DESA", 14, tinggi=22)
    baris(21, f"DESA {DESA}", 14, tinggi=22)
    baris(22, f"KECAMATAN {KECAMATAN}", 14, tinggi=22)

    # kotak tahun
    ws.merge_cells(start_row=25, start_column=4, end_row=25, end_column=5)
    tulis(ws, 25, 4, f"TAHUN {TAHUN}",
          font=Font(name=FONT, size=16, bold=True, color=PUTIH),
          align=TENGAH_ATAS,
          fill=PatternFill("solid", fgColor=HIJAU))
    ws.row_dimensions[25].height = 30

    # ringkasan isi berkas
    baris(29, "REKAPITULASI JUMLAH PEMILIH", 11, color=EMAS, tinggi=20)

    hdr = ["TPS", "LAKI-LAKI", "PEREMPUAN", "JUMLAH"]
    for i, h in enumerate(hdr):
        tulis(ws, 31, 3 + i, h,
              font=Font(name=FONT, size=10, bold=True, color=HIJAU),
              align=TENGAH, border=BOX,
              fill=PatternFill("solid", fgColor=HIJAU_MUDA))
    ws.row_dimensions[31].height = 20

    daftar = [tps] if tps is not None else sorted(per_tps)
    r = 32
    for t in daftar:
        sheet = f"TPS {t:02d}"
        tulis(ws, r, 3, f"TPS {t:02d}", font=Font(name=FONT, size=10),
              align=TENGAH, border=BOX)
        tulis(ws, r, 4, f"='{sheet}'!$B$1", align=TENGAH, border=BOX, fmt="#,##0")
        tulis(ws, r, 5, f"='{sheet}'!$C$1", align=TENGAH, border=BOX, fmt="#,##0")
        tulis(ws, r, 6, f"=SUM(D{r}:E{r})", align=TENGAH, border=BOX, fmt="#,##0",
              font=Font(name=FONT, size=10, bold=True))
        r += 1

    tulis(ws, r, 3, "JUMLAH",
          font=Font(name=FONT, size=10, bold=True, color=PUTIH),
          align=TENGAH, border=BOX, fill=PatternFill("solid", fgColor=HIJAU))
    for c, col in zip(range(4, 7), "DEF"):
        tulis(ws, r, c, f"=SUM({col}32:{col}{r - 1})",
              font=Font(name=FONT, size=10, bold=True, color=PUTIH),
              align=TENGAH, border=BOX, fmt="#,##0",
              fill=PatternFill("solid", fgColor=HIJAU))

    catatan = (
        f"Berkas ini memuat Daftar Pemilih Sementara untuk TPS {tps:02d} saja."
        if tps is not None else
        "Data bersumber dari berkas Coklit DPSHP Desa Sukakarya. "
        "Setiap TPS disajikan pada lembar tersendiri."
    )
    baris(r + 3, catatan, 9, color="595959", bold=False, italic=True, tinggi=18)

    for c in range(2, 8):
        ws.cell(45, c).fill = PatternFill("solid", fgColor=EMAS)
    ws.row_dimensions[45].height = 6

    ws.print_area = f"A1:H{R2 + 1}"
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    return ws


# ============================================================ lembar per TPS
def buat_lembar_tps(wb, tps, baris_data):
    ws = wb.create_sheet(f"TPS {tps:02d}")
    ws.sheet_view.showGridLines = False

    for i, (_, lebar) in enumerate(KOLOM, start=1):
        ws.column_dimensions[get_column_letter(i)].width = lebar

    # --- baris 1 dipakai sebagai sel bantu rekap (disembunyikan) -----------
    n_awal = 10                       # baris data pertama
    n_akhir = n_awal + len(baris_data) - 1
    tulis(ws, 1, 1, "Sel bantu rekapitulasi (jangan dihapus)",
          font=Font(name=FONT, size=8, italic=True, color="808080"))
    tulis(ws, 1, 2, f'=COUNTIF(H{n_awal}:H{n_akhir},"L")', fmt="#,##0",
          font=Font(name=FONT, size=8, color="808080"))
    tulis(ws, 1, 3, f'=COUNTIF(H{n_awal}:H{n_akhir},"P")', fmt="#,##0",
          font=Font(name=FONT, size=8, color="808080"))
    ws.row_dimensions[1].hidden = True

    # --- kop / judul -------------------------------------------------------
    def judul(r, teks, size, bold=True, color=HIJAU, tinggi=None):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOL)
        tulis(ws, r, 1, teks,
              font=Font(name=FONT, size=size, bold=bold, color=color),
              align=TENGAH_ATAS)
        ws.row_dimensions[r].height = tinggi or size * 1.8

    judul(2, "DAFTAR PEMILIH SEMENTARA (DPS)", 16, tinggi=24)
    judul(3, f"PEMILIHAN KEPALA DESA TAHUN {TAHUN}", 13, tinggi=20)
    judul(4, f"DESA {DESA}  -  KECAMATAN {KECAMATAN}  -  KABUPATEN {KABUPATEN}",
          11, tinggi=18)

    # garis emas pemisah
    for c in range(1, NCOL + 1):
        ws.cell(5, c).fill = PatternFill("solid", fgColor=EMAS)
    ws.row_dimensions[5].height = 4

    # --- identitas TPS -----------------------------------------------------
    alamat = sorted({b["alamat"] for b in baris_data})
    rt_rw = sorted({(b["rt"], b["rw"]) for b in baris_data})
    ringkas_rw = sorted({rw for _, rw in rt_rw})

    ws.merge_cells(start_row=7, start_column=1, end_row=7, end_column=4)
    tulis(ws, 7, 1, f"TPS  :  {tps:02d}",
          font=Font(name=FONT, size=12, bold=True, color=HIJAU), align=KIRI)
    ws.merge_cells(start_row=7, start_column=5, end_row=7, end_column=NCOL)
    tulis(ws, 7, 5,
          f"WILAYAH  :  {', '.join(alamat)}   |   RW  :  "
          f"{', '.join(f'{x:02d}' for x in ringkas_rw)}",
          font=Font(name=FONT, size=10), align=KIRI)
    ws.row_dimensions[7].height = 18

    # --- header tabel (2 baris) -------------------------------------------
    hdr_fill = PatternFill("solid", fgColor=HIJAU_MUDA)
    hdr_font = Font(name=FONT, size=9, bold=True, color=HIJAU)

    r1, r2, rnum = 8, 8, 9   # header di baris 8, nomor kolom di baris 9
    for i, (nama, _) in enumerate(KOLOM, start=1):
        tulis(ws, r1, i, nama, font=hdr_font, align=TENGAH,
              border=BOX_TEBAL, fill=hdr_fill)
        tulis(ws, rnum, i, i,
              font=Font(name=FONT, size=8, bold=True, color=HIJAU),
              align=TENGAH, border=BOX, fill=hdr_fill)
    ws.row_dimensions[r1].height = 38
    ws.row_dimensions[rnum].height = 14

    # --- isi tabel ---------------------------------------------------------
    r = n_awal
    for i, b in enumerate(baris_data, start=1):
        ket = DISABILITAS.get(b["disab"], "")
        nilai = [
            i, b["nkk"], b["nik"], b["nama"], b["tempat"], b["tgl"],
            b["kawin"], b["jk"], b["alamat"], b["rt"], b["rw"], ket,
        ]
        strip = PatternFill("solid", fgColor=ABU) if i % 2 == 0 else None
        for c, v in enumerate(nilai, start=1):
            align = KIRI if c in (4, 5, 9) else TENGAH_ATAS
            fmt = None
            if c in (2, 3):
                fmt = "@"
            elif c == 6:
                fmt = "DD-MM-YYYY"
            elif c in (10, 11):
                fmt = "00"
            tulis(ws, r, c, v, font=Font(name=FONT, size=9), align=align,
                  border=BOX, fill=strip, fmt=fmt)
        ws.row_dimensions[r].height = 15
        r += 1

    # --- baris jumlah ------------------------------------------------------
    rj = r
    ws.merge_cells(start_row=rj, start_column=1, end_row=rj, end_column=6)
    tulis(ws, rj, 1, "JUMLAH PEMILIH",
          font=Font(name=FONT, size=10, bold=True, color=HIJAU),
          align=Alignment(horizontal="right", vertical="center"),
          border=BOX_TEBAL, fill=hdr_fill)
    for c in range(7, NCOL + 1):
        tulis(ws, rj, c, None, border=BOX_TEBAL, fill=hdr_fill)

    tulis(ws, rj, 7, f'=COUNTIF(H{n_awal}:H{n_akhir},"L")&" L"',
          font=Font(name=FONT, size=10, bold=True, color=HIJAU),
          align=TENGAH, border=BOX_TEBAL, fill=hdr_fill)
    tulis(ws, rj, 8, f'=COUNTIF(H{n_awal}:H{n_akhir},"P")&" P"',
          font=Font(name=FONT, size=10, bold=True, color=HIJAU),
          align=TENGAH, border=BOX_TEBAL, fill=hdr_fill)
    ws.merge_cells(start_row=rj, start_column=9, end_row=rj, end_column=NCOL)
    tulis(ws, rj, 9, f"=COUNTA(D{n_awal}:D{n_akhir})&\" ORANG\"",
          font=Font(name=FONT, size=10, bold=True, color=HIJAU),
          align=TENGAH, border=BOX_TEBAL, fill=hdr_fill)
    ws.row_dimensions[rj].height = 20

    # --- blok tanda tangan -------------------------------------------------
    rt_ = rj + 3
    ws.merge_cells(start_row=rt_, start_column=8, end_row=rt_, end_column=NCOL)
    tulis(ws, rt_, 8, f"{DESA.title()}, ............................ {TAHUN}",
          font=Font(name=FONT, size=10), align=TENGAH_ATAS)

    ws.merge_cells(start_row=rt_ + 1, start_column=8, end_row=rt_ + 1, end_column=NCOL)
    tulis(ws, rt_ + 1, 8, "PANITIA PEMILIHAN KEPALA DESA",
          font=Font(name=FONT, size=10, bold=True), align=TENGAH_ATAS)
    ws.merge_cells(start_row=rt_ + 2, start_column=8, end_row=rt_ + 2, end_column=NCOL)
    tulis(ws, rt_ + 2, 8, f"DESA {DESA}",
          font=Font(name=FONT, size=10, bold=True), align=TENGAH_ATAS)

    ws.merge_cells(start_row=rt_ + 3, start_column=1, end_row=rt_ + 3, end_column=4)
    tulis(ws, rt_ + 3, 1, "K e t u a,",
          font=Font(name=FONT, size=10), align=TENGAH_ATAS)
    ws.merge_cells(start_row=rt_ + 3, start_column=8, end_row=rt_ + 3, end_column=NCOL)
    tulis(ws, rt_ + 3, 8, "S e k r e t a r i s,",
          font=Font(name=FONT, size=10), align=TENGAH_ATAS)

    for rr in range(rt_ + 4, rt_ + 7):
        ws.row_dimensions[rr].height = 16

    ws.merge_cells(start_row=rt_ + 7, start_column=1, end_row=rt_ + 7, end_column=4)
    tulis(ws, rt_ + 7, 1, "(  ..................................  )",
          font=Font(name=FONT, size=10, bold=True), align=TENGAH_ATAS)
    ws.merge_cells(start_row=rt_ + 7, start_column=8, end_row=rt_ + 7, end_column=NCOL)
    tulis(ws, rt_ + 7, 8, "(  ..................................  )",
          font=Font(name=FONT, size=10, bold=True), align=TENGAH_ATAS)

    # --- pengaturan cetak --------------------------------------------------
    ws.freeze_panes = f"A{n_awal}"
    ws.auto_filter.ref = f"A{rnum}:{LAST_COL}{n_akhir}"
    ws.print_title_rows = f"2:{rnum}"
    ws.print_area = f"A2:{LAST_COL}{rt_ + 7}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = False
    ws.page_margins.left = ws.page_margins.right = 0.3
    ws.page_margins.top = ws.page_margins.bottom = 0.4
    ws.oddFooter.right.text = "Halaman &P dari &N"
    ws.oddFooter.right.size = 8
    ws.oddFooter.left.text = f"DPS Desa {DESA} - TPS {tps:02d}"
    ws.oddFooter.left.size = 8
    return ws


# ============================================================ lembar rekap
def buat_rekap(wb, per_tps):
    ws = wb.create_sheet("REKAPITULASI")
    ws.sheet_view.showGridLines = False
    for col, lebar in zip("ABCDEFG", (4, 12, 26, 13, 13, 13, 13)):
        ws.column_dimensions[col].width = lebar

    def judul(r, teks, size, bold=True, color=HIJAU):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        tulis(ws, r, 1, teks,
              font=Font(name=FONT, size=size, bold=bold, color=color),
              align=TENGAH_ATAS)
        ws.row_dimensions[r].height = size * 1.9

    judul(2, "REKAPITULASI DAFTAR PEMILIH SEMENTARA (DPS)", 15)
    judul(3, f"PEMILIHAN KEPALA DESA {DESA} TAHUN {TAHUN}", 12)
    judul(4, f"KECAMATAN {KECAMATAN} - KABUPATEN {KABUPATEN}", 10)

    hdr = ["NO", "TPS", "WILAYAH (DUKUH / KAMPUNG)", "LAKI-LAKI",
           "PEREMPUAN", "JUMLAH", "DISABILITAS"]
    for i, h in enumerate(hdr, start=1):
        tulis(ws, 6, i, h,
              font=Font(name=FONT, size=10, bold=True, color=HIJAU),
              align=TENGAH, border=BOX_TEBAL,
              fill=PatternFill("solid", fgColor=HIJAU_MUDA))
    ws.row_dimensions[6].height = 30

    r = 7
    for no, tps in enumerate(sorted(per_tps), start=1):
        sheet = f"TPS {tps:02d}"
        rows = per_tps[tps]
        wilayah = ", ".join(sorted({b["alamat"] for b in rows}))
        n_dis = sum(1 for b in rows if b["disab"] in DISABILITAS)
        tulis(ws, r, 1, no, align=TENGAH, border=BOX)
        tulis(ws, r, 2, f"TPS {tps:02d}", align=TENGAH, border=BOX,
              font=Font(name=FONT, size=10, bold=True))
        tulis(ws, r, 3, wilayah, align=KIRI, border=BOX)
        tulis(ws, r, 4, f"='{sheet}'!$B$1", align=TENGAH, border=BOX, fmt="#,##0")
        tulis(ws, r, 5, f"='{sheet}'!$C$1", align=TENGAH, border=BOX, fmt="#,##0")
        tulis(ws, r, 6, f"=SUM(D{r}:E{r})", align=TENGAH, border=BOX, fmt="#,##0",
              font=Font(name=FONT, size=10, bold=True))
        tulis(ws, r, 7, n_dis, align=TENGAH, border=BOX, fmt="#,##0")
        r += 1

    tulis(ws, r, 1, "JUMLAH TOTAL",
          font=Font(name=FONT, size=10, bold=True, color=PUTIH),
          align=TENGAH, border=BOX_TEBAL,
          fill=PatternFill("solid", fgColor=HIJAU))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    for c in range(2, 4):
        tulis(ws, r, c, None, border=BOX_TEBAL,
              fill=PatternFill("solid", fgColor=HIJAU))
    for c, col in zip(range(4, 8), "DEFG"):
        tulis(ws, r, c, f"=SUM({col}7:{col}{r - 1})",
              font=Font(name=FONT, size=10, bold=True, color=PUTIH),
              align=TENGAH, border=BOX_TEBAL, fmt="#,##0",
              fill=PatternFill("solid", fgColor=HIJAU))
    ws.row_dimensions[r].height = 20

    # ---- legenda kode -----------------------------------------------------
    rl = r + 3
    tulis(ws, rl, 1, "KETERANGAN KODE",
          font=Font(name=FONT, size=11, bold=True, color=EMAS), align=KIRI)
    legenda = [
        ("Status Perkawinan", "B = Belum Kawin   |   S = Sudah Kawin   |   P = Pernah Kawin"),
        ("Jenis Kelamin", "L = Laki-laki   |   P = Perempuan"),
        ("Keterangan (Disabilitas)",
         "1 Fisik  |  2 Intelektual  |  3 Mental  |  4 Tuna Wicara  |  "
         "5 Tuna Rungu  |  6 Tuna Netra  |  kosong = tidak disabilitas"),
        ("Syarat usia",
         f"Genap 17 tahun atau sudah/pernah kawin pada "
         f"{TGL_PEMUNGUTAN.strftime('%d %B %Y')} (hari pemungutan suara)"),
    ]
    for i, (k, v) in enumerate(legenda):
        rr = rl + 1 + i
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=2)
        tulis(ws, rr, 1, k, font=Font(name=FONT, size=9, bold=True), align=KIRI)
        ws.merge_cells(start_row=rr, start_column=3, end_row=rr, end_column=7)
        tulis(ws, rr, 3, v, font=Font(name=FONT, size=9), align=KIRI)

    ws.print_area = f"A1:G{rl + len(legenda) + 1}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    return ws


# ============================================================ catatan verifikasi
def buat_catatan(wb, data):
    ws = wb.create_sheet("CATATAN VERIFIKASI")
    ws.sheet_view.showGridLines = False
    for col, lebar in zip("ABCDEF", (5, 9, 20, 30, 14, 46)):
        ws.column_dimensions[col].width = lebar

    ws.merge_cells("A2:F2")
    tulis(ws, 2, 1, "CATATAN HASIL PEMERIKSAAN DATA",
          font=Font(name=FONT, size=14, bold=True, color=HIJAU), align=TENGAH_ATAS)
    ws.merge_cells("A3:F3")
    tulis(ws, 3, 1,
          "Temuan berikut TIDAK diubah/dihapus dari daftar. Baris ini hanya "
          "penanda agar Panitia menelusuri dan memutuskan sendiri statusnya.",
          font=Font(name=FONT, size=9, italic=True, color="595959"), align=TENGAH_ATAS)
    ws.row_dimensions[2].height = 22

    hdr = ["NO", "TPS", "NIK", "NAMA", "JENIS TEMUAN", "URAIAN"]
    for i, h in enumerate(hdr, start=1):
        tulis(ws, 5, i, h,
              font=Font(name=FONT, size=10, bold=True, color=HIJAU),
              align=TENGAH, border=BOX_TEBAL,
              fill=PatternFill("solid", fgColor=HIJAU_MUDA))
    ws.row_dimensions[5].height = 20

    temuan = []

    # 1. usia di bawah 17 tahun dan belum kawin
    for b in data:
        u = usia_pada(b["tgl"])
        if u is not None and u < 17:
            temuan.append((
                b, "Usia < 17 tahun",
                f"Lahir {b['tgl'].strftime('%d-%m-%Y')} (usia {u} th pada "
                f"{TGL_PEMUNGUTAN.strftime('%d-%m-%Y')}), status kawin "
                f"'{b['kawin']}'. Memenuhi syarat hanya bila sudah/pernah kawin.",
            ))

    # 2. NIK ganda
    idx = defaultdict(list)
    for b in data:
        idx[b["nik"]].append(b)
    for nik, grup in sorted(idx.items()):
        if len(grup) > 1:
            tps_list = ", ".join(f"TPS {g['tps']:02d}" for g in grup)
            for b in grup:
                temuan.append((
                    b, "NIK ganda",
                    f"NIK muncul {len(grup)} kali ({tps_list}). "
                    f"Nama tercatat: {' / '.join(g['nama'] for g in grup)}.",
                ))

    temuan.sort(key=lambda t: (t[0]["tps"], t[1], t[0]["nama"]))

    r = 6
    for no, (b, jenis, uraian) in enumerate(temuan, start=1):
        strip = PatternFill("solid", fgColor=ABU) if no % 2 == 0 else None
        for c, v in enumerate(
            [no, f"TPS {b['tps']:02d}", b["nik"], b["nama"], jenis, uraian], start=1
        ):
            tulis(ws, r, c, v, font=Font(name=FONT, size=9),
                  align=KIRI if c in (3, 4, 6) else TENGAH_ATAS,
                  border=BOX, fill=strip, fmt="@" if c == 3 else None)
        ws.row_dimensions[r].height = 26
        r += 1

    rr = r + 2
    tulis(ws, rr, 1, "Perbaikan yang sudah diterapkan pada berkas ini:",
          font=Font(name=FONT, size=10, bold=True, color=EMAS), align=KIRI)
    perbaikan = [
        "Ditambahkan halaman COVER dan kop/judul resmi pada setiap lembar.",
        "Data dipisah menjadi satu lembar kerja per TPS (TPS 01 s.d. TPS 11).",
        "Nomor urut pemilih dihitung ulang dari 1 pada masing-masing TPS.",
        "Spasi ganda/berlebih pada kolom alamat dirapikan (mis. 'KP. KUDA-KUDA ').",
        "Penulisan kolom E-KTP diseragamkan menjadi huruf kapital.",
        "Kode disabilitas diterjemahkan ke teks pada kolom KETERANGAN.",
        "NKK dan NIK diformat sebagai teks agar tidak berubah jadi notasi ilmiah.",
        "Tanggal lahir diseragamkan dengan format DD-MM-YYYY.",
        "Pengaturan cetak: A4 landscape, judul berulang tiap halaman, nomor halaman.",
        "Lembar REKAPITULASI memakai rumus, sehingga ikut berubah bila data diedit.",
        "Lembar sumber 'Olah Data' tetap disertakan tanpa perubahan.",
    ]
    for i, p in enumerate(perbaikan):
        ws.merge_cells(start_row=rr + 1 + i, start_column=1,
                       end_row=rr + 1 + i, end_column=6)
        tulis(ws, rr + 1 + i, 1, f"{i + 1}.  {p}",
              font=Font(name=FONT, size=9), align=KIRI)

    ws.print_area = f"A1:F{rr + len(perbaikan) + 1}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = False
    return ws


# ============================================================ main
def kelompokkan(data):
    per_tps = defaultdict(list)
    for b in data:
        per_tps[b["tps"]].append(b)
    return per_tps


def main():
    data = baca_sumber()
    per_tps = kelompokkan(data)

    # bangun di atas berkas asli agar sheet "Olah Data" tetap utuh
    shutil.copy(SRC, OUT)
    wb = openpyxl.load_workbook(OUT)

    n = perbaiki_olah_data(wb)
    print(f"rumus DATEDIF diperbaiki: {n}")

    buat_cover(wb, per_tps)
    buat_rekap(wb, per_tps)
    for tps in sorted(per_tps):
        buat_lembar_tps(wb, tps, per_tps[tps])
    buat_catatan(wb, data)

    # urutan tab: COVER, REKAPITULASI, TPS 01..11, CATATAN, Olah Data
    urutan = (["COVER", "REKAPITULASI"]
              + [f"TPS {t:02d}" for t in sorted(per_tps)]
              + ["CATATAN VERIFIKASI", "Olah Data"])
    wb._sheets = [wb[n] for n in urutan]
    wb.active = 0
    wb["COVER"].sheet_view.tabSelected = True
    for n in urutan[1:]:
        wb[n].sheet_view.tabSelected = False

    wb.save(OUT)
    print(f"tersimpan: {OUT}")
    print(f"total pemilih: {len(data)} | jumlah TPS: {len(per_tps)}")
    for t in sorted(per_tps):
        print(f"  TPS {t:02d}: {len(per_tps[t])}")


if __name__ == "__main__":
    main()
